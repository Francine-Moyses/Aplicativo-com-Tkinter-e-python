import os
import pandas as pd
import sys
import os.path
from openpyxl import load_workbook
import openpyxl
import fsspec
from tkinter import *

def GerarPasta():
    diretório = '/'
    diretório_pasta = '/'

    arquivo = vnome.get()

    wb = load_workbook(diretório_pasta + arquivo + '.xlsx')
    ws = wb.active

    n = int(vlinhas.get())

    for row in range(2, n):
        var1 = ws.cell(row=row, column=1).value
        var2 = ws.cell(row=row, column=2).value
        var3 = ws.cell(row=row, column=3).value
        var4 = ws.cell(row=row, column=4).value
        var1_1 = str(var1)
        var2_1 = str(var2)
        var3_1 = str(var3)
        var4_1 = str(var4)
        if not os.path.exists(diretório + var1_1 + '/' + var2_1 + '_' + var3_1 + '/A' + var3_1 + var4_1):
            os.makedirs(diretório + var1_1 + '/' + var2_1 + '_' + var3_1 + '/A' + var3_1 + var4_1)
            print("Pasta criada")

appGP=Tk()
appGP.title("Aplicativo | Criar Pasta")
appGP.geometry("320x350")
appGP.configure(background="#14540d")


txt1=Label(appGP,text="Nome do arquivo",background="#14540d",foreground="#ffffff")
txt1.place(x=10,y=30,width=120,height=30)

vnome=Entry(appGP)
vnome.place(x=22,y=60,width=250,height=20)

txt2=Label(appGP,text="Quantidade de linhas",background="#14540d",foreground="#ffffff")
txt2.place(x=20,y=120,width=120,height=30)

vlinhas=Entry(appGP)
vlinhas.place(x=22,y=150,width=100,height=20)

btn1=Button(appGP, text="Criar Pastas",background="#ffffff",foreground="#000000", command=GerarPasta)
btn1.place(x=80,y=250,width=140,height=30)

appGP.mainloop()
